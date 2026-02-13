"""Import Brysbaert GPT familiarity data into the frequencies table.

Reads the full 417K-word familiarity dataset, matches against lemmas in our
lexicon, computes rarity tiers, and populates the frequencies table. Lemmas
not found in any dataset get rarity='unusual' with NULL familiarity.

Usage:
    python import_familiarity.py
"""
import sqlite3
from pathlib import Path

import openpyxl

from utils import (
    LEXICON_V2,
    FAMILIARITY_FULL_XLSX,
    FAMILIARITY_COMMON_THRESHOLD,
    FAMILIARITY_UNUSUAL_THRESHOLD,
    ZIPF_COMMON_THRESHOLD,
    ZIPF_UNUSUAL_THRESHOLD,
)


def compute_rarity(familiarity: float | None, zipf: float | None) -> str:
    """Compute rarity tier from familiarity (primary) or Zipf (fallback)."""
    if familiarity is not None:
        if familiarity >= FAMILIARITY_COMMON_THRESHOLD:
            return "common"
        elif familiarity >= FAMILIARITY_UNUSUAL_THRESHOLD:
            return "unusual"
        else:
            return "rare"
    elif zipf is not None:
        if zipf >= ZIPF_COMMON_THRESHOLD:
            return "common"
        elif zipf >= ZIPF_UNUSUAL_THRESHOLD:
            return "unusual"
        else:
            return "rare"
    else:
        return "unusual"


def load_familiarity(xlsx_path: Path) -> dict[str, tuple[float, int, float | None]]:
    """Load familiarity data from the full Brysbaert xlsx.

    Returns dict mapping lowercase word -> (GPT_Fam_probs, GPT_Fam_dominant, Multilex_Zipf or None).
    Skips null words, multi-word expressions (Type != 'W'), and duplicates (keeps first).
    """
    print(f"Loading familiarity data from {xlsx_path.name}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Find column indices from header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    word_idx = headers.index("Word")
    fam_dom_idx = headers.index("Dom_Pos")
    fam_probs_idx = headers.index("Freq_Dom")
    multilex_idx = headers.index("MultLex_Percent")
    entry_type_idx = headers.index("EntryType")

    data: dict[str, tuple[float, int, float | None]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        word = row[word_idx]
        fam_dom = row[fam_dom_idx]
        fam_probs = row[fam_probs_idx]
        multilex = row[multilex_idx]
        entry_type = row[entry_type_idx]

        # Skip nulls and non-single-words
        if word is None or not isinstance(word, str):
            continue
        if entry_type != "W":
            continue

        key = word.lower().strip()
        if key in data:
            continue  # keep first occurrence

        fam_probs_val = float(fam_probs) if fam_probs is not None else None
        fam_dom_val = int(fam_dom) if fam_dom is not None else None

        # Multilex column can be 'NA' string or numeric
        multilex_val = None
        if multilex is not None and multilex != "NA":
            try:
                multilex_val = float(multilex)
            except (ValueError, TypeError):
                pass

        if fam_probs_val is not None:
            data[key] = (fam_probs_val, fam_dom_val, multilex_val)

    wb.close()
    print(f"  Loaded {len(data)} single-word entries")
    return data


def get_all_lemmas(conn: sqlite3.Connection) -> set[str]:
    """Get all distinct lemmas from the lexicon."""
    cursor = conn.execute("SELECT DISTINCT lemma FROM lemmas")
    return {row[0] for row in cursor}


def import_familiarity(conn: sqlite3.Connection, familiarity_data: dict) -> dict[str, int]:
    """Import familiarity data into frequencies table.

    Returns stats dict with counts.
    """
    lemmas = get_all_lemmas(conn)
    print(f"  {len(lemmas)} distinct lemmas in lexicon")

    # Clear existing data
    conn.execute("DELETE FROM frequencies")

    rows = []
    stats = {"matched": 0, "matched_hyphen": 0, "unmatched": 0}

    for lemma in lemmas:
        fam_probs = None
        fam_dom = None
        multilex_zipf = None
        source = None

        # Try exact match (case-insensitive — both sides already lowercase)
        if lemma in familiarity_data:
            fam_probs, fam_dom, multilex_zipf = familiarity_data[lemma]
            source = "brysbaert"
            stats["matched"] += 1
        else:
            # Try stripping hyphens
            dehyphenated = lemma.replace("-", " ")
            if dehyphenated != lemma and dehyphenated in familiarity_data:
                fam_probs, fam_dom, multilex_zipf = familiarity_data[dehyphenated]
                source = "brysbaert"
                stats["matched_hyphen"] += 1
            else:
                stats["unmatched"] += 1

        rarity = compute_rarity(fam_probs, multilex_zipf)

        rows.append((
            lemma,
            fam_probs,
            fam_dom,
            multilex_zipf,  # zipf column — Multilex Zipf for now, SUBTLEX later
            None,           # frequency — will be filled by SUBTLEX import
            rarity,
            source,
        ))

    conn.executemany(
        """INSERT OR REPLACE INTO frequencies
           (lemma, familiarity, familiarity_dominant, zipf, frequency, rarity, source)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        rows,
    )
    conn.commit()

    return stats


def main():
    if not FAMILIARITY_FULL_XLSX.exists():
        raise FileNotFoundError(f"Familiarity data not found: {FAMILIARITY_FULL_XLSX}")
    if not LEXICON_V2.exists():
        raise FileNotFoundError(f"Lexicon DB not found: {LEXICON_V2}")

    familiarity_data = load_familiarity(FAMILIARITY_FULL_XLSX)

    conn = sqlite3.connect(LEXICON_V2)
    try:
        stats = import_familiarity(conn, familiarity_data)
    finally:
        conn.close()

    total = stats["matched"] + stats["matched_hyphen"] + stats["unmatched"]
    print(f"\nImport complete:")
    print(f"  Exact match:   {stats['matched']} ({stats['matched'] / total * 100:.1f}%)")
    print(f"  Hyphen match:  {stats['matched_hyphen']} ({stats['matched_hyphen'] / total * 100:.1f}%)")
    print(f"  Unmatched:     {stats['unmatched']} ({stats['unmatched'] / total * 100:.1f}%)")
    print(f"  Total:         {total}")


if __name__ == "__main__":
    main()
