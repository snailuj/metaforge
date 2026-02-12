"""Import SUBTLEX-UK frequency data as a second pass over the frequencies table.

Backfills zipf and frequency columns for lemmas that have SUBTLEX-UK data.
Updates source to reflect combined provenance.
Must run AFTER import_familiarity.py.

Usage:
    python import_subtlex.py
"""
import sqlite3
from pathlib import Path

import openpyxl

from utils import (
    LEXICON_V2,
    SUBTLEX_FLEMMAS_XLSX,
    ZIPF_COMMON_THRESHOLD,
    ZIPF_UNUSUAL_THRESHOLD,
)


def load_subtlex_flemmas(xlsx_path: Path) -> dict[str, tuple[float, int]]:
    """Load SUBTLEX-UK flemma data.

    Returns dict mapping lowercase flemma -> (Zipf, frequency_count).
    """
    print(f"Loading SUBTLEX-UK flemmas from {xlsx_path.name}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Find column indices from header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    flemma_idx = headers.index("flemma")
    zipf_idx = headers.index("Zipf")
    freq_idx = headers.index("lemmafreqs_combined")

    data: dict[str, tuple[float, int]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        flemma = row[flemma_idx]
        if flemma is None or not isinstance(flemma, str):
            continue

        key = flemma.lower().strip()
        if key in data:
            continue

        try:
            zipf_val = float(row[zipf_idx]) if row[zipf_idx] is not None else None
        except (ValueError, TypeError):
            continue
        try:
            raw_freq = row[freq_idx]
            if raw_freq is not None:
                # Handle comma-formatted numbers (e.g. "2,0" or "1,234")
                freq_val = int(str(raw_freq).replace(",", "")) if raw_freq else None
            else:
                freq_val = None
        except (ValueError, TypeError):
            freq_val = None

        if zipf_val is not None:
            data[key] = (zipf_val, freq_val)

    wb.close()
    print(f"  Loaded {len(data)} flemma entries")
    return data


def backfill_subtlex(conn: sqlite3.Connection, subtlex_data: dict) -> dict[str, int]:
    """Backfill SUBTLEX data into existing frequencies rows.

    For rows that already have familiarity (from Brysbaert), only update zipf/frequency
    if the current zipf is NULL. For rows with no familiarity data, also recompute
    rarity using Zipf fallback thresholds.

    Returns stats dict.
    """
    stats = {"updated_zipf": 0, "rarity_recomputed": 0, "no_match": 0}

    # Get all current frequency rows
    rows = conn.execute(
        "SELECT lemma, familiarity, zipf, source FROM frequencies"
    ).fetchall()

    for lemma, familiarity, current_zipf, current_source in rows:
        if lemma not in subtlex_data:
            stats["no_match"] += 1
            continue

        subtlex_zipf, subtlex_freq = subtlex_data[lemma]

        # Update zipf and frequency from SUBTLEX (prefer SUBTLEX over Multilex for Zipf)
        new_source = current_source
        if current_source == "brysbaert":
            new_source = "brysbaert+subtlex"
        elif current_source is None:
            new_source = "subtlex"

        # Recompute rarity if familiarity is NULL (Zipf-only fallback)
        new_rarity = None
        if familiarity is None:
            if subtlex_zipf >= ZIPF_COMMON_THRESHOLD:
                new_rarity = "common"
            elif subtlex_zipf >= ZIPF_UNUSUAL_THRESHOLD:
                new_rarity = "unusual"
            else:
                new_rarity = "rare"
            stats["rarity_recomputed"] += 1

        if new_rarity is not None:
            conn.execute(
                """UPDATE frequencies
                   SET zipf = ?, frequency = ?, source = ?, rarity = ?
                   WHERE lemma = ?""",
                (subtlex_zipf, subtlex_freq, new_source, new_rarity, lemma),
            )
        else:
            conn.execute(
                """UPDATE frequencies
                   SET zipf = ?, frequency = ?, source = ?
                   WHERE lemma = ?""",
                (subtlex_zipf, subtlex_freq, new_source, lemma),
            )
        stats["updated_zipf"] += 1

    conn.commit()
    return stats


def main():
    if not SUBTLEX_FLEMMAS_XLSX.exists():
        raise FileNotFoundError(f"SUBTLEX-UK flemmas not found: {SUBTLEX_FLEMMAS_XLSX}")
    if not LEXICON_V2.exists():
        raise FileNotFoundError(f"Lexicon DB not found: {LEXICON_V2}")

    subtlex_data = load_subtlex_flemmas(SUBTLEX_FLEMMAS_XLSX)

    conn = sqlite3.connect(LEXICON_V2)
    try:
        stats = backfill_subtlex(conn, subtlex_data)
    finally:
        conn.close()

    print(f"\nSUBTLEX backfill complete:")
    print(f"  Zipf updated:       {stats['updated_zipf']}")
    print(f"  Rarity recomputed:  {stats['rarity_recomputed']}")
    print(f"  No SUBTLEX match:   {stats['no_match']}")


if __name__ == "__main__":
    main()
